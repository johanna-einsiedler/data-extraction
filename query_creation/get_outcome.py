"""Extract outcome labels from the latest true-labels spreadsheet into a JSON map."""

import json
import re
from pathlib import Path

import openpyxl

# Path setup -----------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "true_labels"

THREE_DIGIT_PATTERN = re.compile(r"^\d{3}$")


def clean_value(value):
    """Strip boilerplate prefixes and whitespace from header cells."""
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"(?i)^outcome:\s*", "", text)


def extract_outcomes(workbook_path: Path) -> dict:
    """
    Convert the first worksheet of the outcome workbook into a doc_id -> outcome map.

    The function assumes row 2 contains human-readable headers and row 3
    contains three-digit document identifiers.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[2]]
    doc_ids = [cell.value for cell in ws[3]]

    outcomes = {}
    for header, doc_id in zip(headers, doc_ids):
        doc_id_str = "" if doc_id is None else str(doc_id).strip()
        if doc_id_str and THREE_DIGIT_PATTERN.match(doc_id_str):
            outcomes[doc_id_str] = clean_value(header)

    return outcomes


def latest_outcome_workbook(data_dir: Path) -> Path:
    """Return the most recently modified .xlsx file in the given directory."""
    xlsx_files = list(data_dir.glob("*.xlsx"))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in {data_dir}")
    return max(xlsx_files, key=lambda f: f.stat().st_ctime)


def main() -> None:
    """Entry point when executing the module as a script."""
    workbook_path = latest_outcome_workbook(DATA_DIR)
    outcomes = extract_outcomes(workbook_path)

    output_path = workbook_path.parent / "outcomes.json"
    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(outcomes, fh, indent=2, ensure_ascii=False)

    print(f"Saved JSON to: {output_path}")


if __name__ == "__main__":
    main()
